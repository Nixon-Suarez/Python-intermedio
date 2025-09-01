import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime

def cargar_inventario():
    try:
        wb = openpyxl.load_workbook('Datos_Iniciales.xlsx')
    except FileNotFoundError:
        print("Error no se encontró el archivo 'Datos_Iniciales.xlsx'.")
        return None
    
    ws = wb.active
    ws.title = "Inventario"

    #Aplicar formato a los encabezados
    for col in range(1, ws.max_column + 1):
        celda = ws.cell(row=1, column=col)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        celda.alignment = Alignment(horizontal="center", vertical="center")

    return wb, ws

def actualizar_precios(ws, porcentaje):
    """Actualizar precios en la hoja de inventario"""
    for row in range(2, ws.max_row + 1):
        precio_actual = float(ws.cell(row=row, column=4).value)
        nuevo_precio = precio_actual * (1 + porcentaje / 100)
        ws.cell(row=row, column=4).value = round(nuevo_precio, 2)

def automatizacion_inventario():
    """Funcion principal para automatizar el inventario"""
    resultado = cargar_inventario()  # Primero vemos qué devuelve
    
    if resultado is None:
        return
    
    wb, ws = resultado  # Si no es None, ahora sí podemos desempaquetar

    # Actualizamos precios
    actualizar_precios(ws, 5) # Aumentar precios en un 5%
    
    wb.save('inventario_V2.xlsx')
    print("\nProceso de automatizacion_inventario ejecutado correctamente")


if __name__ == "__main__":
    automatizacion_inventario()
