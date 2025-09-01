import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def crear_excel_inicial():
    # Crear un nuevo libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Iniciales"

    # Definir los encabezados de las columnas
    encabezados = [
        "ID",
        "Producto",
        "Categoria",
        "Precio",
        "Stock"
    ]

    datos = [
        [1, "Producto A", "Categoria 1", 10.0, 100],
        [2, "Producto B", "Categoria 2", 20.0, 200],
        [3, "Producto C", "Categoria 1", 30.0, 300],
        [4, "Producto D", "Categoria 3", 40.0, 400],
        [5, "Producto E", "Categoria 2", 50.0, 500]
    ]

    # Escribir los encabezados 
    for col_num, encabezado in enumerate(encabezados, 1): 
        # los valores encabezados(ID, Producto...) se guardan en la variable encabezados
        # la enumeracion se asigna a la variable col_num en otras palabras A cada encabezado se le asigna un número de columna 
        celda = ws.cell(row=1, column=col_num)
        celda.value = encabezado
        celda.font = Font(bold=True) # se le asigna un fondo
        celda.alignment = Alignment(horizontal="center") # se alinea el texto al centro
        celda.fill = PatternFill(start_color="b8cce4", end_color="b8cce4", fill_type="solid") # se le asigna un color de fondo a la celda
        
    # Escribir los datos
    for row_num, fila in enumerate(datos, 2):  # Comenzar desde la fila 2 ya que el uno es elcabezado
        # por cada lista(fila) en datos se asigna un número de fila
        for col_num, valor in enumerate(fila, 1):  # Comenzar desde la columna 1
            # por cada valor en fila se asigna un número de columna
            celda = ws.cell(row=row_num, column=col_num)
            celda.value = valor
            celda.alignment = Alignment(horizontal="center")

    # ajustar ancho del las columnas
    for col in range(1, len(encabezados) + 1): # +1 porque el rango no incluye el último número
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    # Guardar el archivo Excel
    wb.save("Datos_Iniciales.xlsx")
    print("Archivo 'Datos_Iniciales.xlsx' creado exitosamente.")

if __name__ == "__main__":
    crear_excel_inicial()